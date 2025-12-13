from __future__ import annotations

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import Shipment


def build_shipments_workbook_bytes(*, date_from: date | None, date_to: date | None) -> bytes:
    """Build an Excel file (XLSX) with shipments for a date range.

    Filters by Message.date (date part). If dates are None, exports all.
    """
    qs = Shipment.objects.select_related('message__channel').order_by('-message__date')

    if date_from:
        qs = qs.filter(message__date__date__gte=date_from)
    if date_to:
        qs = qs.filter(message__date__date__lte=date_to)

    wb = Workbook()
    ws = wb.active
    ws.title = "Shipments"

    headers = [
        'channel_id', 'channel_title',
        'message_id', 'date',
        'origin', 'destination',
        'cargo_type', 'truck_type',
        'payment_type', 'phone',
        'text',
    ]
    ws.append(headers)

    for sh in qs:
        msg = sh.message
        ch = msg.channel
        ws.append([
            ch.channel_id,
            ch.title or "",
            msg.message_id,
            msg.date.isoformat() if msg.date else "",
            sh.origin or "",
            sh.destination or "",
            sh.cargo_type or "",
            sh.truck_type or "",
            sh.payment_type or "",
            sh.phone or "",
            (msg.text or "")[:5000],
        ])

    # Reasonable widths
    widths = {
        'A': 14,
        'B': 24,
        'C': 12,
        'D': 22,
        'E': 18,
        'F': 18,
        'G': 22,
        'H': 12,
        'I': 12,
        'J': 16,
        'K': 60,
    }
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = widths.get(col_letter, 18)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
