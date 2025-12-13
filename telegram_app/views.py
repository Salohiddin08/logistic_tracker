import asyncio

from django.conf import settings
from django.contrib.auth import logout
from django.db.models import Count
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.core.paginator import Paginator
from django.views.decorators.http import require_POST
from asgiref.sync import async_to_sync

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import TelegramSession, Channel, Message, Shipment
from .telethon_client import get_client, get_channels, get_messages
from .utils import save_messages_json, parse_shipment_text
from .bot_service import send_export_now
from telethon import TelegramClient
from telethon.sessions import StringSession
from telethon.errors import SessionPasswordNeededError


def home_view(request):
    """Home page: show dashboard if session exists, else redirect to Telegram login."""
    if TelegramSession.objects.last():
        return redirect('dashboard')
    return redirect('telegram_phone_login')


def dashboard_view(request):
    """Simple dashboard for today."""
    today = timezone.localdate()

    shipments = Shipment.objects.select_related('message__channel').filter(message__date__date=today)

    total_today = shipments.count()

    top_origins = (
        shipments.values('origin')
        .annotate(total=Count('id'))
        .order_by('-total')[:10]
    )
    top_destinations = (
        shipments.values('destination')
        .annotate(total=Count('id'))
        .order_by('-total')[:10]
    )
    top_payments = (
        shipments.values('payment_type')
        .annotate(total=Count('id'))
        .order_by('-total')[:10]
    )
    top_cargo = (
        shipments.values('cargo_type')
        .annotate(total=Count('id'))
        .order_by('-total')[:10]
    )

    sent = request.GET.get('sent')
    error = request.GET.get('err')

    context = {
        'today': today,
        'total_today': total_today,
        'top_origins': top_origins,
        'top_destinations': top_destinations,
        'top_payments': top_payments,
        'top_cargo': top_cargo,
        'sent': sent,
        'error': error,
    }
    return render(request, 'dashboard.html', context)


@require_POST
def bot_export_view(request):
    """Send Excel export to admin chat via Telegram bot."""
    try:
        days = int(request.POST.get('days') or 1)
    except Exception:
        days = 1

    try:
        async_to_sync(send_export_now)(days=days)
        return redirect(f"/dashboard/?sent=1")
    except Exception as exc:
        return redirect(f"/dashboard/?err={str(exc)}")


def add_session(request):
    # Oddiy forma orqali API ID / HASH / StringSession qabul qilamiz
    if request.method == 'POST':
        api_id = request.POST.get('api_id')
        api_hash = request.POST.get('api_hash')
        string_session = request.POST.get('string_session')

        TelegramSession.objects.create(
            api_id=api_id,
            api_hash=api_hash,
            string_session=string_session
        )
        return redirect('channels')

    default_api_id = getattr(settings, 'TG_API_ID', '')
    default_api_hash = getattr(settings, 'TG_API_HASH', '')

    context = {
        'default_api_id': default_api_id,
        'default_api_hash': default_api_hash,
    }
    return render(request, 'add_session.html', context)


def _get_tg_credentials():
    api_id = getattr(settings, 'TG_API_ID', None)
    api_hash = getattr(settings, 'TG_API_HASH', None)

    if not api_id or not api_hash:
        raise ValueError("TG_API_ID / TG_API_HASH .env faylda to'ldirilmagan")

    # Telethon api_id int bo'lishini kutadi
    try:
        api_id = int(api_id)
    except Exception as exc:
        raise ValueError("TG_API_ID raqam (int) bo'lishi kerak") from exc

    return api_id, api_hash


async def _start_phone_login(phone: str):
    """Telefon raqamni qabul qilib, kod yuboradi va vaqtinchalik sessiyani qaytaradi."""
    api_id, api_hash = _get_tg_credentials()
    client = TelegramClient(StringSession(), api_id, api_hash)
    await client.connect()
    sent = await client.send_code_request(phone)
    temp_session = client.session.save()
    return temp_session, sent.phone_code_hash


async def _complete_phone_login(temp_session: str, phone: str, code: str, password: str | None, phone_code_hash: str | None):
    """Kod va (bo'lsa) 2-bosqich parol bilan login qilish.

    Telethon oqimi:
      1) send_code_request -> phone_code_hash
      2) sign_in(phone, code, phone_code_hash=...)
      3) Agar 2FA yoqilgan bo'lsa: SessionPasswordNeededError -> sign_in(password=...)
    """
    api_id, api_hash = _get_tg_credentials()
    client = TelegramClient(StringSession(temp_session), api_id, api_hash)
    await client.connect()

    # 1-bosqich: SMS kodi bilan kirish.
    # Agar akkauntda 2FA yoqilgan bo'lsa, bu yerda SessionPasswordNeededError chiqadi.
    try:
        await client.sign_in(phone=phone, code=code, phone_code_hash=phone_code_hash)
    except SessionPasswordNeededError:
        # Two-step verification yoqilgan
        if not password:
            raise ValueError("Telegram akkauntingizda 2-bosqichli parol (Two‑step verification) yoqilgan. Parolni kiriting.")
        await client.sign_in(password=password)

    # Agar 2FA talab qilinmagan bo'lsa, lekin password berilgan bo'lsa ham muammo emas.
    # (Telethon ba'zida password sign_in ni qayta chaqirishni talab qilmaydi)

    return client.session.save()


def telegram_phone_login(request):
    """1-bosqich: telefon raqamni kiritish.

    Telefon kiritilgach, Telethon orqali kod yuboriladi va keyingi bosqichga yo'naltiriladi.
    """
    error = None

    if request.method == 'POST':
        phone = request.POST.get('phone')
        if phone:
            try:
                temp_session, phone_code_hash = asyncio.run(_start_phone_login(phone))
                request.session['tg_phone'] = phone
                request.session['tg_temp_session'] = temp_session
                request.session['tg_phone_code_hash'] = phone_code_hash
                return redirect('telegram_phone_code')
            except Exception as exc:
                error = str(exc)

    return render(request, 'telegram_login_phone.html', {'error': error})


def telegram_phone_code(request):
    """2-bosqich: SMS kodi va (bo'lsa) 2-bosqich parol.

    Yakunida StringSession yaratiladi va TelegramSession jadvaliga saqlanadi.
    """
    phone = request.session.get('tg_phone')
    temp_session = request.session.get('tg_temp_session')
    phone_code_hash = request.session.get('tg_phone_code_hash')
    if not phone or not temp_session:
        return redirect('telegram_phone_login')

    error = None
    if request.method == 'POST':
        code = request.POST.get('code')
        password = request.POST.get('password') or None
        if code:
            try:
                string_session = asyncio.run(
                    _complete_phone_login(temp_session, phone, code, password, phone_code_hash)
                )

                # Yangi sessiyani bazaga yozamiz va vaqtinchalik sessiyani tozalaymiz
                api_id = getattr(settings, 'TG_API_ID', '')
                api_hash = getattr(settings, 'TG_API_HASH', '')
                TelegramSession.objects.create(
                    api_id=api_id,
                    api_hash=api_hash,
                    string_session=string_session,
                )
                request.session.pop('tg_phone', None)
                request.session.pop('tg_temp_session', None)
                request.session.pop('tg_phone_code_hash', None)
                return redirect('channels')
            except Exception as exc:
                # Masalan, kod noto'g'ri bo'lsa yoki boshqa Telethon xatosi bo'lsa,
                # 500 o'rniga oddiy xabar ko'rsatamiz.
                error = str(exc)

    context = {
        'phone': phone,
        'error': error,
    }
    return render(request, 'telegram_login_code.html', context)


def channels_view(request):
    session = TelegramSession.objects.last()
    if not session:
        # Agar Telegram sessiya bo'lmasa, birinchi navbatda telefon-login sahifasiga yuboramiz
        return redirect('telegram_phone_login')

    async def run():
        client = await get_client(session.api_id, session.api_hash, session.string_session)
        return await get_channels(client)

    try:
        channels = asyncio.run(run())
    except Exception as exc:
        # Debug rejimda detail ko'rinadi; aks holda ham foydali xabar chiqadi
        return render(
            request,
            'error.html',
            {
                'title': 'Telegram ulanish xatosi',
                'message': "Kanallarni olishda xatolik. Session eskirgan bo'lishi yoki TG_API_ID/TG_API_HASH noto'g'ri bo'lishi mumkin.",
                'detail': str(exc),
            },
            status=500,
        )

    return render(request, 'channels.html', {'channels': channels})


def fetch_messages_view(request, channel_id):
    session = TelegramSession.objects.last()
    if not session:
        return redirect('add_session')

    async def run():
        client = await get_client(session.api_id, session.api_hash, session.string_session)
        return await get_messages(client, channel_id, limit=100)

    messages = asyncio.run(run())

    channel_obj, _ = Channel.objects.get_or_create(channel_id=channel_id)

    for m in messages:
        msg_obj, _ = Message.objects.get_or_create(
            channel=channel_obj,
            message_id=m.id,
            defaults={
                'sender_id': getattr(m.from_id, 'user_id', None),
                'sender_name': getattr(m.sender, 'username', None) if m.sender else None,
                'text': m.message,
                'date': m.date,
            },
        )

        # Har bir xabar matnidan yuk ma'lumotlarini parslash
        parsed = parse_shipment_text(m.message or "")
        Shipment.objects.update_or_create(
            message=msg_obj,
            defaults={
                'origin': parsed.get('origin'),
                'destination': parsed.get('destination'),
                'cargo_type': parsed.get('cargo_type'),
                'truck_type': parsed.get('truck_type'),
                'payment_type': parsed.get('payment_type'),
                'phone': parsed.get('phone'),
            },
        )

    return redirect('channel_stats', channel_id=channel_id)


def saved_messages_view(request):
    """Saqlangan xabarlarni sana bo'yicha filtrlash bilan ko'rsatish."""
    messages = Message.objects.select_related('channel').order_by('-date')

    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    parsed_from = parse_date(date_from) if date_from else None
    parsed_to = parse_date(date_to) if date_to else None

    if parsed_from:
        messages = messages.filter(date__date__gte=parsed_from)
    if parsed_to:
        messages = messages.filter(date__date__lte=parsed_to)

    # Har bir sahifada 20 tadan xabar ko'rsatamiz
    paginator = Paginator(messages, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'messages': page_obj.object_list,
        'page_obj': page_obj,
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'messages.html', context)


def _get_filtered_shipments(request, channel_id):
    """Yordamchi funksiya: kanal va sana bo'yicha Shipment querysetini qaytaradi."""
    shipments = Shipment.objects.filter(message__channel__channel_id=channel_id)

    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    parsed_from = parse_date(date_from) if date_from else None
    parsed_to = parse_date(date_to) if date_to else None

    if parsed_from:
        shipments = shipments.filter(message__date__date__gte=parsed_from)
    if parsed_to:
        shipments = shipments.filter(message__date__date__lte=parsed_to)

    return shipments, date_from, date_to



def channel_stats_view(request, channel_id):
    """Tanlangan kanal bo'yicha yuk statistikasi (sana filtri bilan)."""
    shipments, date_from, date_to = _get_filtered_shipments(request, channel_id)

    # A → B yo'nalishlar jadvali uchun pagination
    route_qs = (
        shipments
        .values('origin', 'destination')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    route_paginator = Paginator(route_qs, 20)
    route_page_number = request.GET.get('route_page')
    route_page_obj = route_paginator.get_page(route_page_number)
    route_stats = route_page_obj.object_list

    # Yuk turlari jadvali uchun pagination
    cargo_qs = (
        shipments
        .values('cargo_type')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    cargo_paginator = Paginator(cargo_qs, 20)
    cargo_page_number = request.GET.get('cargo_page')
    cargo_page_obj = cargo_paginator.get_page(cargo_page_number)
    cargo_stats = cargo_page_obj.object_list

    # Transport turlari jadvali uchun pagination
    truck_qs = (
        shipments
        .values('truck_type')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    truck_paginator = Paginator(truck_qs, 20)
    truck_page_number = request.GET.get('truck_page')
    truck_page_obj = truck_paginator.get_page(truck_page_number)
    truck_stats = truck_page_obj.object_list

    # To'lov turlari jadvali uchun pagination
    payment_qs = (
        shipments
        .values('payment_type')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    payment_paginator = Paginator(payment_qs, 20)
    payment_page_number = request.GET.get('payment_page')
    payment_page_obj = payment_paginator.get_page(payment_page_number)
    payment_stats = payment_page_obj.object_list

    context = {
        'channel_id': channel_id,
        'total_shipments': shipments.count(),
'route_stats': route_stats,
        'route_page_obj': route_page_obj,
        'cargo_stats': cargo_stats,
        'truck_stats': truck_stats,
        'payment_stats': payment_stats,
        'cargo_page_obj': cargo_page_obj,
        'truck_page_obj': truck_page_obj,
        'payment_page_obj': payment_page_obj,
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'stats.html', context)


def channel_stats_excel(request, channel_id):
    """Tanlangan kanal bo'yicha yuk ma'lumotlarini (sana filtri bilan) Excel (.xlsx) formatida yuklab berish."""
    shipments, date_from, date_to = _get_filtered_shipments(request, channel_id)

    filename_parts = [f"channel_{channel_id}"]
    if date_from:
        filename_parts.append(f"from_{date_from}")
    if date_to:
        filename_parts.append(f"to_{date_to}")
    filename_base = "_".join(filename_parts)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Shipments"

    headers = [
        'channel_id', 'channel_title', 'message_id', 'date',
        'origin', 'destination', 'cargo_type', 'truck_type',
        'payment_type', 'phone',
    ]
    ws.append(headers)

    for shipment in shipments.select_related('message__channel'):
        msg = shipment.message
        ch = msg.channel
        ws.append([
            ch.channel_id,
            ch.title or "",
            msg.message_id,
            msg.date.isoformat() if msg.date else "",
            shipment.origin or "",
            shipment.destination or "",
            shipment.cargo_type or "",
            shipment.truck_type or "",
            shipment.payment_type or "",
            shipment.phone or "",
        ])

    # Avtomatik ustun kengliklari
    for idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = 18

    wb.save(response)
    return response


def channel_phones_view(request, channel_id):
    """Kanal bo'yicha unikal telefon raqamlar ro'yxati (sana bo'yicha filtrlash bilan).

    Telefonlarga o'xshash raqamlar (uzunroq, + bilan boshlanadigan) va qisqa ID/raqamlarni
    alohida ro'yxatlarga ajratamiz.
    """
    shipments, date_from, date_to = _get_filtered_shipments(request, channel_id)

    raw_stats = (
        shipments
        .exclude(phone__isnull=True)
        .exclude(phone__exact="")
        .values('phone')
        .annotate(total=Count('id'))
        .order_by('-total')
    )

    phone_stats = []  # haqiqiy telefonlar
    id_stats = []     # ID / boshqa qisqa raqamlar

    for item in raw_stats:
        phone = item['phone'] or ""
        digits_only = ''.join(ch for ch in phone if ch.isdigit())
        # Oddiy heuristika: + bilan boshlangan yoki uzunroq raqamlar – telefon sifatida
        if phone.startswith('+') or len(digits_only) >= 9:
            phone_stats.append(item)
        else:
            id_stats.append(item)

    context = {
        'channel_id': channel_id,
        'phone_stats': phone_stats,
        'id_stats': id_stats,
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'phones.html', context)


def channel_phones_excel(request, channel_id):
    """Unikal telefon raqamlarini Excel (.xlsx) formatida yuklash."""
    shipments, date_from, date_to = _get_filtered_shipments(request, channel_id)

    phone_stats = (
        shipments
        .exclude(phone__isnull=True)
        .exclude(phone__exact="")
        .values('phone')
        .annotate(total=Count('id'))
        .order_by('-total')
    )

    filename_parts = [f"channel_{channel_id}_phones"]
    if date_from:
        filename_parts.append(f"from_{date_from}")
    if date_to:
        filename_parts.append(f"to_{date_to}")
    filename_base = "_".join(filename_parts)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Phones"

    headers = ['phone', 'total_shipments']
    ws.append(headers)

    for item in phone_stats:
        ws.append([
            item['phone'],
            item['total'],
        ])

    for idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = 20

    wb.save(response)
    return response


def channel_phone_messages_view(request, channel_id):
    """Muayyan telefon raqami bo'yicha xabarlar ro'yxati."""
    shipments, date_from, date_to = _get_filtered_shipments(request, channel_id)

    phone = request.GET.get('phone') or None
    if phone:
        shipments = shipments.filter(phone=phone)

    shipments = shipments.select_related('message__channel').order_by('-message__date')

    context = {
        'channel_id': channel_id,
        'phone': phone,
        'origin': None,
        'destination': None,
        'cargo_type': None,
        'truck_type': None,
        'payment_type': None,
        'date_from': date_from,
        'date_to': date_to,
        'shipments': shipments,
    }
    return render(request, 'phone_messages.html', context)


def channel_route_messages_view(request, channel_id):
    """Muayyan yo'nalish (origin/destination) bo'yicha xabarlar ro'yxati."""
    shipments, date_from, date_to = _get_filtered_shipments(request, channel_id)

    origin = request.GET.get('origin') or None
    destination = request.GET.get('destination') or None

    if origin:
        shipments = shipments.filter(origin=origin)
    if destination:
        shipments = shipments.filter(destination=destination)

    shipments = shipments.select_related('message__channel').order_by('-message__date')

    context = {
        'channel_id': channel_id,
        'origin': origin,
        'destination': destination,
        'cargo_type': None,
        'date_from': date_from,
        'date_to': date_to,
        'shipments': shipments,
    }
    return render(request, 'route_messages.html', context)


def channel_cargo_messages_view(request, channel_id):
    """Muayyan yuk turi bo'yicha xabarlar ro'yxati."""
    shipments, date_from, date_to = _get_filtered_shipments(request, channel_id)

    cargo_type = request.GET.get('cargo_type') or None
    if cargo_type:
        shipments = shipments.filter(cargo_type=cargo_type)

    shipments = shipments.select_related('message__channel').order_by('-message__date')

    context = {
        'channel_id': channel_id,
        'origin': None,
        'destination': None,
        'cargo_type': cargo_type,
        'date_from': date_from,
        'date_to': date_to,
        'shipments': shipments,
    }
    return render(request, 'route_messages.html', context)


def channel_truck_messages_view(request, channel_id):
    """Muayyan transport (truck_type) bo'yicha xabarlar ro'yxati."""
    shipments, date_from, date_to = _get_filtered_shipments(request, channel_id)

    truck_type = request.GET.get('truck_type') or None
    if truck_type:
        shipments = shipments.filter(truck_type=truck_type)

    shipments = shipments.select_related('message__channel').order_by('-message__date')

    context = {
        'channel_id': channel_id,
        'origin': None,
        'destination': None,
        'cargo_type': None,
        'truck_type': truck_type,
        'payment_type': None,
        'date_from': date_from,
        'date_to': date_to,
        'shipments': shipments,
    }
    return render(request, 'route_messages.html', context)


def channel_payment_messages_view(request, channel_id):
    """Muayyan to'lov turi (payment_type) bo'yicha xabarlar ro'yxati."""
    shipments, date_from, date_to = _get_filtered_shipments(request, channel_id)

    payment_type = request.GET.get('payment_type') or None
    if payment_type:
        shipments = shipments.filter(payment_type=payment_type)

    shipments = shipments.select_related('message__channel').order_by('-message__date')

    context = {
        'channel_id': channel_id,
        'origin': None,
        'destination': None,
        'cargo_type': None,
        'truck_type': None,
        'payment_type': payment_type,
        'date_from': date_from,
        'date_to': date_to,
        'shipments': shipments,
    }
    return render(request, 'route_messages.html', context)


def export_json(request):
    save_messages_json()
    return HttpResponse("JSON file created successfully!")


def logout_view(request):
    """Oddiy GET orqali ham chiqishni qo'llab-quvvatlaydigan logout."""
    logout(request)
    return redirect('login')
